from __future__ import annotations

from dataclasses import replace
from datetime import datetime
from pathlib import Path
from typing import List, Optional

from PySide6.QtCore import Qt, QTimer
from PySide6.QtGui import QDoubleValidator, QFont
from PySide6.QtWidgets import (
    QApplication,
    QFrame,
    QFileDialog,
    QGridLayout,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QMainWindow,
    QMessageBox,
    QPushButton,
    QScrollArea,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
    QHeaderView,
    QSizePolicy,
)
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from src.core.calc import Settings, calc_item, calc_quick, calc_rub_per_coin_buyer
from src.core.config import GoodsItem, load_config, load_goods, new_goods_item, save_config, save_goods
from src.services.rate_service import fetch_rate
from src.ui.settings_dialog import SettingsDialog

APP_QSS = """
* {
    font-family: "Segoe UI", "Inter", sans-serif;
    font-size: 12.5pt;
}
QMainWindow {
    background-color: #0f141a;
    color: #e6e6e6;
}
QFrame#Card {
    background-color: #161b22;
    border: 1px solid #242b36;
    border-radius: 10px;
}
QLabel#CardTitle {
    font-size: 12.5pt;
    font-weight: 600;
    color: #cfd6df;
}
QLineEdit, QComboBox {
    background-color: #0f1216;
    border: 1px solid #303744;
    border-radius: 6px;
    padding: 6px 8px;
    color: #e6e6e6;
    min-height: 34px;
}
QPushButton {
    background-color: #2d6cdf;
    border: none;
    border-radius: 8px;
    padding: 6px 12px;
    color: #ffffff;
    min-height: 36px;
}
QPushButton:hover { background-color: #3a7bff; }
QPushButton:disabled { background-color: #3a3f48; color: #9aa3ad; }
QTableWidget {
    background-color: #141821;
    border: 1px solid #242b36;
    gridline-color: #242b36;
    color: #e6e6e6;
}
QHeaderView::section {
    background-color: #1b2028;
    padding: 6px;
    border: 1px solid #242b36;
    color: #cfd6df;
}
#StatusBadge {
    padding: 2px 8px;
    border-radius: 10px;
    background-color: #26303c;
}
#HintLabel {
    color: #ffb5b5;
    font-size: 10.5pt;
}
"""


class MainWindow(QMainWindow):
    def __init__(self) -> None:
        super().__init__()
        self.setWindowTitle("KapManiak — L2 Trade Helper")

        self.config = load_config()
        self.goods: List[GoodsItem] = load_goods()

        self._build_ui()
        self._load_config_to_fields()
        self._refresh_quick_calc()
        self._refresh_goods_table()

        self.rate_timer = QTimer(self)
        self.rate_timer.setInterval(10 * 60 * 1000)
        self.rate_timer.timeout.connect(self.update_rate)
        self.rate_timer.start()
        self.update_rate()

    def _build_ui(self) -> None:
        central = QWidget()
        main_layout = QVBoxLayout(central)
        main_layout.setContentsMargins(16, 16, 16, 16)
        main_layout.setSpacing(12)

        main_layout.addWidget(self._build_top_bar())

        content_layout = QHBoxLayout()
        content_layout.setSpacing(14)

        left_scroll = QScrollArea()
        left_scroll.setWidgetResizable(True)
        left_scroll.setFrameShape(QScrollArea.NoFrame)
        left_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        left_scroll.setFixedWidth(420)

        left_container = QWidget()
        left_layout = QVBoxLayout(left_container)
        left_layout.setContentsMargins(0, 0, 0, 0)
        left_layout.setSpacing(12)

        left_layout.addWidget(self._build_params_card())
        left_layout.addStretch(1)

        left_scroll.setWidget(left_container)

        content_layout.addWidget(left_scroll)
        content_layout.addWidget(self._build_goods_card(), 1)

        main_layout.addLayout(content_layout)
        self.setCentralWidget(central)

    def _build_top_bar(self) -> QWidget:
        container = QWidget()
        container.setFixedHeight(56)
        layout = QHBoxLayout(container)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(10)

        self.rate_label = QLabel("Курс USDT: —")
        self.status_label = QLabel("NO")
        self.status_label.setObjectName("StatusBadge")
        self.status_label.setAlignment(Qt.AlignCenter)

        refresh_button = QPushButton("Обновить")
        refresh_button.clicked.connect(self.update_rate)

        settings_button = QPushButton("⚙")
        settings_button.setFixedWidth(44)
        settings_button.clicked.connect(self.open_settings)

        layout.addWidget(self.rate_label)
        layout.addWidget(self.status_label)
        layout.addStretch(1)
        layout.addWidget(refresh_button)
        layout.addWidget(settings_button)
        return container

    def _build_card(self, title: str, header_extra: Optional[QWidget] = None) -> tuple[QFrame, QVBoxLayout]:
        card = QFrame()
        card.setObjectName("Card")
        card_layout = QVBoxLayout(card)
        card_layout.setContentsMargins(16, 16, 16, 16)
        card_layout.setSpacing(12)

        header_layout = QHBoxLayout()
        title_label = QLabel(title)
        title_label.setObjectName("CardTitle")
        title_font = QFont()
        title_font.setBold(True)
        title_label.setFont(title_font)
        header_layout.addWidget(title_label)
        header_layout.addStretch(1)
        if header_extra is not None:
            header_layout.addWidget(header_extra)
        card_layout.addLayout(header_layout)
        return card, card_layout

    def _build_params_card(self) -> QFrame:
        card, layout = self._build_card("Расчёт")

        self.coin_to_adena_input = self._make_number_input("1 монета = X адены")
        self.rub_per_1kk_input = self._make_number_input("1кк адены = ₽")
        self.coins_qty_input = self._make_number_input("Кол-во монет")
        self.coins_qty_input.textChanged.connect(self._refresh_quick_calc)

        self.rub_per_coin_base_label = self._make_value_label()
        self.rub_per_coin_sbp_label = self._make_value_label()
        self.sum_base_label = self._make_value_label()
        self.sum_sbp_label = self._make_value_label()
        self.sum_withdraw_rub_label = self._make_value_label()
        self.sum_withdraw_usdt_label = self._make_value_label()

        form_layout = QGridLayout()
        form_layout.setHorizontalSpacing(10)
        form_layout.setVerticalSpacing(10)
        form_layout.addWidget(QLabel("1 монета = адены"), 0, 0)
        form_layout.addWidget(self.coin_to_adena_input, 0, 1)
        form_layout.addWidget(QLabel("1кк адены = ₽ (FP)"), 1, 0)
        form_layout.addWidget(self.rub_per_1kk_input, 1, 1)
        form_layout.addWidget(QLabel("Кол-во монет"), 2, 0)
        form_layout.addWidget(self.coins_qty_input, 2, 1)
        form_layout.addWidget(QLabel("1 монета (База) ="), 3, 0)
        form_layout.addWidget(self.rub_per_coin_base_label, 3, 1)
        form_layout.addWidget(QLabel("1 монета (СБП) ="), 4, 0)
        form_layout.addWidget(self.rub_per_coin_sbp_label, 4, 1)
        form_layout.addWidget(QLabel("Сумма (База) ₽"), 5, 0)
        form_layout.addWidget(self.sum_base_label, 5, 1)
        form_layout.addWidget(QLabel("Сумма (СБП) ₽"), 6, 0)
        form_layout.addWidget(self.sum_sbp_label, 6, 1)
        form_layout.addWidget(QLabel("К выводу ₽"), 7, 0)
        form_layout.addWidget(self.sum_withdraw_rub_label, 7, 1)
        form_layout.addWidget(QLabel("К получению USDT"), 8, 0)
        form_layout.addWidget(self.sum_withdraw_usdt_label, 8, 1)
        form_layout.setColumnStretch(1, 1)

        layout.addLayout(form_layout)
        self.params_save_timer = QTimer(self)
        self.params_save_timer.setSingleShot(True)
        self.params_save_timer.timeout.connect(self.save_params)
        self.coin_to_adena_input.textChanged.connect(self._schedule_save_params)
        self.rub_per_1kk_input.textChanged.connect(self._schedule_save_params)
        return card

    def _build_goods_card(self) -> QFrame:
        card, layout = self._build_card("Товары")

        self.item_name_input = QLineEdit()
        self.item_name_input.setPlaceholderText("Название товара (необязательно)")
        self.item_price_input = self._make_number_input("Цена в монетах")

        form_layout = QGridLayout()
        form_layout.setHorizontalSpacing(10)
        form_layout.setVerticalSpacing(10)
        form_layout.addWidget(QLabel("Товар"), 0, 0)
        form_layout.addWidget(self.item_name_input, 0, 1)
        form_layout.addWidget(QLabel("Цена (монеты)"), 1, 0)
        form_layout.addWidget(self.item_price_input, 1, 1)
        form_layout.setColumnStretch(1, 1)

        add_button = QPushButton("Добавить")
        add_button.clicked.connect(self.add_goods)

        remove_button = QPushButton("Удалить")
        remove_button.clicked.connect(self.remove_selected_goods)

        clear_button = QPushButton("Очистить")
        clear_button.clicked.connect(self.clear_goods)

        export_button = QPushButton("Сохранить")
        export_button.clicked.connect(self.export_goods)

        for button in (add_button, remove_button, clear_button, export_button):
            button.setMinimumWidth(120)
            button.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)

        buttons_layout = QHBoxLayout()
        buttons_layout.setSpacing(10)
        buttons_layout.addWidget(add_button)
        buttons_layout.addWidget(remove_button)
        buttons_layout.addWidget(clear_button)
        buttons_layout.addWidget(export_button)
        buttons_layout.addStretch(1)

        self.goods_hint = QLabel("")
        self.goods_hint.setObjectName("HintLabel")
        self.goods_hint.setWordWrap(True)

        self.goods_table = QTableWidget(0, 7)
        self.goods_table.setHorizontalHeaderLabels(
            [
                "Товар",
                "Цена (монеты)",
                "База ₽",
                "Карта RU ₽",
                "СБП QR ₽",
                "К выводу ₽",
                "К получению USDT",
            ]
        )
        self.goods_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.goods_table.setSelectionMode(QTableWidget.SingleSelection)
        self.goods_table.verticalHeader().setVisible(False)
        self.goods_table.setWordWrap(False)
        self.goods_table.setTextElideMode(Qt.ElideMiddle)
        header = self.goods_table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.Stretch)
        header.setStretchLastSection(False)
        header.setMinimumSectionSize(140)
        self.goods_table.setColumnWidth(0, 200)
        self.goods_table.setColumnWidth(1, 140)

        layout.addLayout(form_layout)
        layout.addLayout(buttons_layout)
        layout.addWidget(self.goods_hint)
        layout.addWidget(self.goods_table)
        return card

    def _make_number_input(self, placeholder: str) -> QLineEdit:
        field = QLineEdit()
        field.setPlaceholderText(placeholder)
        field.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
        validator = QDoubleValidator(0.0, 1_000_000_000.0, 6, field)
        validator.setNotation(QDoubleValidator.StandardNotation)
        field.setValidator(validator)
        return field

    def _make_value_label(self) -> QLabel:
        label = QLabel("—")
        label.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
        return label

    def _load_config_to_fields(self) -> None:
        self.coin_to_adena_input.setText(_format_number(self.config.coin_to_adena))
        self.rub_per_1kk_input.setText(_format_number(self.config.rub_per_1kk_buyer))

    def save_params(self) -> None:
        coin_to_adena = _parse_positive_float(self.coin_to_adena_input.text())
        rub_per_1kk = _parse_positive_float(self.rub_per_1kk_input.text())
        self.config = replace(self.config, coin_to_adena=coin_to_adena, rub_per_1kk_buyer=rub_per_1kk)
        save_config(self.config)
        self._refresh_quick_calc()
        self._refresh_goods_table()
        self._persist_goods()

    def open_settings(self) -> None:
        dialog = SettingsDialog(
            self.config.funpay_fee,
            self.config.sbp_fee_effective,
            self.config.k_card_ru,
            self.config.k_sbp_qr,
            self.config.withdraw_fee_pct,
            self.config.withdraw_fee_min_rub,
            self.config.withdraw_rate_rub_per_usdt or self.config.rub_per_usdt,
            self,
        )
        if dialog.exec() == dialog.Accepted:
            funpay_fee = dialog.parse_percent(dialog.funpay_fee_input.text())
            sbp_fee_effective = dialog.parse_percent(dialog.sbp_fee_effective_input.text())
            k_card_ru = dialog.parse_number(dialog.k_card_ru_input.text())
            k_sbp_qr = dialog.parse_number(dialog.k_sbp_qr_input.text())
            withdraw_fee_pct = dialog.parse_percent(dialog.withdraw_fee_pct_input.text())
            withdraw_fee_min_rub = dialog.parse_number(dialog.withdraw_fee_min_rub_input.text())
            withdraw_rate_rub_per_usdt = dialog.parse_number(dialog.withdraw_rate_rub_per_usdt_input.text())
            self.config = replace(
                self.config,
                funpay_fee=funpay_fee if funpay_fee is not None else self.config.funpay_fee,
                sbp_fee_effective=(
                    sbp_fee_effective if sbp_fee_effective is not None else self.config.sbp_fee_effective
                ),
                k_card_ru=k_card_ru if k_card_ru is not None else self.config.k_card_ru,
                k_sbp_qr=k_sbp_qr if k_sbp_qr is not None else self.config.k_sbp_qr,
                withdraw_fee_pct=(
                    withdraw_fee_pct if withdraw_fee_pct is not None else self.config.withdraw_fee_pct
                ),
                withdraw_fee_min_rub=(
                    withdraw_fee_min_rub
                    if withdraw_fee_min_rub is not None
                    else self.config.withdraw_fee_min_rub
                ),
                withdraw_rate_rub_per_usdt=(
                    withdraw_rate_rub_per_usdt
                    if withdraw_rate_rub_per_usdt is not None
                    else self.config.withdraw_rate_rub_per_usdt
                ),
            )
            save_config(self.config)
            self._refresh_quick_calc()
            self._refresh_goods_table()
            self._persist_goods()

    def update_rate(self) -> None:
        self.status_label.setText("UPD")
        self.status_label.setStyleSheet("background-color: #3a3f48; padding: 2px 8px; border-radius: 10px;")
        QApplication.processEvents()
        result = fetch_rate()
        if result.rate is not None:
            updated_withdraw_rate = self.config.withdraw_rate_rub_per_usdt
            if updated_withdraw_rate is None or updated_withdraw_rate == self.config.rub_per_usdt:
                updated_withdraw_rate = result.rate
            self.config = replace(
                self.config,
                rub_per_usdt=result.rate,
                withdraw_rate_rub_per_usdt=updated_withdraw_rate,
            )
            save_config(self.config)
            self.status_label.setText("OK")
            self.status_label.setStyleSheet("background-color: #1f6f50; padding: 2px 8px; border-radius: 10px;")
        else:
            self.status_label.setText("NO")
            self.status_label.setStyleSheet("background-color: #6b2b2b; padding: 2px 8px; border-radius: 10px;")
        self.rate_label.setText(f"Курс USDT: {_format_rub(self.config.rub_per_usdt, suffix=' ₽')}")
        self._refresh_quick_calc()
        self._refresh_goods_table()
        self._persist_goods()

    def _refresh_quick_calc(self) -> None:
        settings = self._settings()
        rub_per_coin_buyer = calc_rub_per_coin_buyer(settings)
        rub_per_coin_me = (
            rub_per_coin_buyer * (1 - settings.funpay_fee) if rub_per_coin_buyer is not None else None
        )
        sbp_multiplier = settings.k_sbp_qr or None
        rub_per_coin_sbp = rub_per_coin_me * sbp_multiplier if rub_per_coin_me is not None else None
        self.rub_per_coin_base_label.setText(_format_rub(rub_per_coin_me))
        self.rub_per_coin_sbp_label.setText(_format_rub(rub_per_coin_sbp))

        coins_qty = _parse_positive_float(self.coins_qty_input.text())
        quick = calc_quick(settings, coins_qty, None)
        self.sum_base_label.setText(_format_rub_total(quick.base_rub))
        self.sum_sbp_label.setText(_format_rub_total(quick.sbp_rub))
        self.sum_withdraw_rub_label.setText(_format_rub_total(quick.withdraw_amount_rub))
        self.sum_withdraw_usdt_label.setText(_format_usdt_range(quick.withdraw_usdt))

    def add_goods(self) -> None:
        price_coins = _parse_positive_float(self.item_price_input.text())
        if price_coins is None:
            self.goods_hint.setText("Введите цену товара в монетах.")
            return
        if calc_rub_per_coin_buyer(self._settings()) is None:
            self.goods_hint.setText("Заполните курс монеты к адене и ₽ за 1кк.")
            return
        item = new_goods_item(self.item_name_input.text(), price_coins)
        self.goods.append(item)
        self._persist_goods()
        self.item_name_input.clear()
        self.item_price_input.clear()
        self.goods_hint.setText("")
        self._refresh_goods_table()

    def remove_selected_goods(self) -> None:
        selection = self.goods_table.selectionModel().selectedRows()
        if not selection:
            return
        index = selection[0].row()
        if 0 <= index < len(self.goods):
            self.goods.pop(index)
            self._persist_goods()
            self._refresh_goods_table()

    def clear_goods(self) -> None:
        if not self.goods:
            return
        self.goods = []
        save_goods(self.goods)
        self._refresh_goods_table()

    def _refresh_goods_table(self) -> None:
        self.goods_table.setRowCount(len(self.goods))
        settings = self._settings()
        for row_index, item in enumerate(self.goods):
            calc = calc_item(settings, item.price_coins)
            values = [
                item.name,
                _format_coins(item.price_coins),
                _format_rub(calc.base_rub),
                _format_rub(calc.card_rub),
                _format_rub(calc.sbp_rub),
                _format_rub(calc.withdraw_amount_rub),
                _format_usdt(calc.withdraw_usdt),
            ]
            for col, text in enumerate(values):
                cell = QTableWidgetItem(text)
                if col > 0:
                    cell.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                self.goods_table.setItem(row_index, col, cell)
        self.goods_table.resizeRowsToContents()

    def export_goods(self) -> None:
        filename, _ = QFileDialog.getSaveFileName(
            self,
            "Сохранить как…",
            "kapmaniak_items.xlsx",
            "Excel Files (*.xlsx)",
        )
        if not filename:
            return
        path = Path(filename)
        if path.suffix.lower() != ".xlsx":
            path = path.with_suffix(".xlsx")

        settings = self._settings()
        workbook = Workbook()
        items_sheet = workbook.active
        items_sheet.title = "Items"

        headers = [
            "Товар",
            "Цена (монеты)",
            "База ₽",
            "Карта RU ₽",
            "СБП QR ₽",
            "К выводу ₽",
            "К получению USDT",
        ]
        items_sheet.append(headers)
        header_font = Font(bold=True)
        for col_index in range(1, len(headers) + 1):
            cell = items_sheet.cell(row=1, column=col_index)
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for item in self.goods:
            calc = calc_item(settings, item.price_coins)
            items_sheet.append(
                [
                    item.name,
                    item.price_coins,
                    calc.base_rub,
                    calc.card_rub,
                    calc.sbp_rub,
                    calc.withdraw_amount_rub,
                    calc.withdraw_usdt,
                ]
            )

        number_align = Alignment(horizontal="right")
        coin_format = "# ##0.00##"
        rub_format = "# ##0.00"
        usdt_format = "# ##0.00##"
        for row in items_sheet.iter_rows(min_row=2, min_col=2, max_col=7):
            for col_offset, cell in enumerate(row, start=2):
                if cell.value is None:
                    continue
                cell.alignment = number_align
                if col_offset == 2:
                    cell.number_format = coin_format
                elif col_offset == 7:
                    cell.number_format = usdt_format
                else:
                    cell.number_format = rub_format

        items_sheet.freeze_panes = "A2"
        _autosize_columns(items_sheet)

        meta_sheet = workbook.create_sheet("Meta")
        meta_headers = ["Параметр", "Значение"]
        meta_sheet.append(meta_headers)
        for col_index in range(1, len(meta_headers) + 1):
            cell = meta_sheet.cell(row=1, column=col_index)
            cell.font = header_font

        meta_rows = [
            ("timestamp", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("USDT_RUB rate", settings.rub_per_usdt),
            ("withdraw_rate ₽/USDT", settings.withdraw_rate_rub_per_usdt or settings.rub_per_usdt),
            ("withdraw_fee_pct", settings.withdraw_fee_pct),
            ("withdraw_fee_min", settings.withdraw_fee_min_rub),
            ("k_sbp_qr", settings.k_sbp_qr),
            ("k_card_ru", settings.k_card_ru),
            ("1 монета = адены", settings.coin_to_adena),
            ("1кк адены = ₽ (FP)", settings.rub_per_1kk_buyer),
        ]
        for key, value in meta_rows:
            meta_sheet.append([key, value])

        for row in meta_sheet.iter_rows(min_row=2, min_col=2, max_col=2):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.alignment = number_align

        _autosize_columns(meta_sheet)
        workbook.save(path)
        QMessageBox.information(self, "Экспорт", f"Сохранено: {path}")

    def _settings(self) -> Settings:
        return Settings(
            coin_to_adena=self.config.coin_to_adena,
            rub_per_1kk_buyer=self.config.rub_per_1kk_buyer,
            funpay_fee=self.config.funpay_fee,
            sbp_fee_effective=self.config.sbp_fee_effective,
            k_card_ru=self.config.k_card_ru,
            k_sbp_qr=self.config.k_sbp_qr,
            withdraw_fee_pct=self.config.withdraw_fee_pct,
            withdraw_fee_min_rub=self.config.withdraw_fee_min_rub,
            withdraw_rate_rub_per_usdt=self.config.withdraw_rate_rub_per_usdt,
            rub_per_usdt=self.config.rub_per_usdt,
        )

    def _schedule_save_params(self) -> None:
        self.params_save_timer.start(300)

    def _persist_goods(self) -> None:
        settings = self._settings()
        for item in self.goods:
            calc = calc_item(settings, item.price_coins)
            item.base_rub = calc.base_rub
            item.card_rub = calc.card_rub
            item.sbp_rub = calc.sbp_rub
            item.withdraw_amount_rub = calc.withdraw_amount_rub
            item.withdraw_usdt = calc.withdraw_usdt
        save_goods(self.goods)


def _parse_positive_float(text: str) -> Optional[float]:
    normalized = text.replace(",", ".").strip()
    if not normalized:
        return None
    try:
        value = float(normalized)
    except ValueError:
        return None
    return value if value > 0 else None


def _format_number(value: Optional[float]) -> str:
    if value is None:
        return ""
    return f"{value:.6f}".rstrip("0").rstrip(".")


def _format_coins(value: Optional[float]) -> str:
    if value is None:
        return "—"
    formatted = f"{value:,.6f}"
    whole, _, frac = formatted.partition(".")
    frac = frac.rstrip("0")
    if len(frac) < 2:
        frac = frac.ljust(2, "0")
    return f"{whole}.{frac}".replace(",", " ")


def _format_rub(value: Optional[float], suffix: str = " ₽") -> str:
    if value is None:
        return "—"
    if abs(value) < 1:
        formatted = f"{value:,.4f}"
        whole, _, frac = formatted.partition(".")
        frac = frac.rstrip("0")
        if len(frac) < 2:
            frac = frac.ljust(2, "0")
        formatted = f"{whole}.{frac}"
    else:
        formatted = f"{value:,.2f}"
    return f"{formatted}{suffix}".replace(",", " ")


def _format_rub_total(value: Optional[float], suffix: str = " ₽") -> str:
    if value is None:
        return "—"
    formatted = f"{value:,.2f}".replace(",", " ")
    return f"{formatted}{suffix}"


def _format_usdt(value: Optional[float]) -> str:
    if value is None:
        return "—"
    return f"{value:,.4f} USDT".replace(",", " ")


def _format_usdt_range(value: Optional[float]) -> str:
    if value is None:
        return "—"
    formatted = f"{value:,.4f}".replace(",", " ")
    whole, _, frac = formatted.partition(".")
    frac = frac.rstrip("0")
    if len(frac) < 2:
        frac = frac.ljust(2, "0")
    formatted = f"{whole}.{frac}"
    return f"{formatted} USDT"


def _format_percent(value: Optional[float]) -> str:
    if value is None:
        return "—"
    return f"{value * 100:.2f}%"


def _autosize_columns(sheet: object) -> None:
    for column_cells in sheet.columns:
        max_length = 0
        column = column_cells[0].column
        for cell in column_cells:
            value = cell.value
            if value is None:
                continue
            length = len(str(value))
            if length > max_length:
                max_length = length
        adjusted = max(10, min(max_length + 2, 40))
        sheet.column_dimensions[get_column_letter(column)].width = adjusted
